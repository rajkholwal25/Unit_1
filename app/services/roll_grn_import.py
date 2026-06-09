"""Parse GRN batch Excel uploads and import new raw-material rows."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

def _openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError(
            'Excel support requires openpyxl. Install it: pip install openpyxl'
        ) from exc
    return Workbook, load_workbook, get_column_letter

from app.extensions import db
from app.models.roll_grn import RollGrnEntry
from app.services.roll_grn_service import (
    _decimal_or_none,
    _max_grn_sequence,
    format_grn_number,
    get_roll_grn_by_supplier_roll,
)

# Header aliases — more specific labels first within each field group.
_COLUMN_ALIASES: dict[str, list[str]] = {
    'supplier_name': ['supplier name', 'supplier', 'vendor name', 'vendor'],
    'supplier_roll_number': [
        'supplier roll number', 'roll number', 'roll no.', 'roll no', 'roll #', 'roll id',
    ],
    'film_type': ['film type', 'material type', 'film', 'material'],
    'coating': ['chemical coating', 'coating', 'chemical'],
    'width_mm': ['width (mm)', 'width mm', 'width'],
    'thickness_mic': ['thickness (mic)', 'thickness mic', 'thickness (micron)', 'thickness'],
    'length_mtr': ['length (mtr)', 'length mtr', 'length (m)', 'length'],
    'gross_weight_kg': ['gross weight (kg)', 'gross weight kg', 'gross weight', 'gross wt', 'gross'],
    'net_weight_kg': ['net weight (kg)', 'net weight kg', 'net weight', 'net wt', 'net'],
    'core_weight_kg': ['core weight (kg)', 'core weight kg', 'core weight', 'core wt', 'core'],
}

_TEMPLATE_HEADERS = [
    ('Supplier Name', 'ABC Films Ltd'),
    ('Roll Number', 'SUP-R-1024'),
    ('Film Type', 'PET'),
    ('Chemical Coating', 'AC'),
    ('Width (mm)', 1200),
    ('Thickness (mic)', 12),
    ('Length (mtr)', 5000),
    ('Gross Weight (kg)', 450.5),
    ('Net weight (kg)', 440),
    ('Core weight (kg)', 10.5),
]


def _norm_header(val: Any) -> str:
    s = str(val or '').strip().lower()
    s = re.sub(r'\s+', ' ', s)
    return s


def _map_headers(header_row: list[Any]) -> dict[str, int]:
    normalized = [_norm_header(c) for c in header_row]
    mapping: dict[str, int] = {}
    used_cols: set[int] = set()

    for field, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            for idx, cell in enumerate(normalized):
                if idx in used_cols:
                    continue
                if cell == alias or cell.replace(' ', '') == alias.replace(' ', ''):
                    mapping[field] = idx
                    used_cols.add(idx)
                    break
            if field in mapping:
                break

    return mapping


def _cell_val(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    if isinstance(val, str):
        return val.strip()
    return val


def _parse_rows_from_workbook(wb) -> tuple[dict[str, int], list[tuple[Any, ...]], int]:
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('Excel file is empty.')

    header_idx = None
    col_map: dict[str, int] = {}
    for i, row in enumerate(rows):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        candidate = _map_headers(list(row))
        if len(candidate) >= 4:
            header_idx = i
            col_map = candidate
            break

    if header_idx is None:
        raise ValueError(
            'Could not find a header row. Use the template columns: '
            'Supplier Name, Roll Number, Film Type, Chemical Coating, Width (mm), …'
        )

    required = [
        'supplier_name', 'supplier_roll_number', 'film_type', 'coating',
        'width_mm', 'thickness_mic', 'length_mtr', 'gross_weight_kg', 'net_weight_kg',
    ]
    missing = [k for k in required if k not in col_map]
    if missing:
        labels = {
            'supplier_name': 'Supplier Name',
            'supplier_roll_number': 'Roll Number',
            'film_type': 'Film Type',
            'coating': 'Chemical Coating',
            'width_mm': 'Width (mm)',
            'thickness_mic': 'Thickness (mic)',
            'length_mtr': 'Length (mtr)',
            'gross_weight_kg': 'Gross Weight (kg)',
            'net_weight_kg': 'Net weight (kg)',
        }
        raise ValueError('Missing columns: ' + ', '.join(labels.get(k, k) for k in missing))

    data_rows = []
    for row in rows[header_idx + 1:]:
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        data_rows.append(row)

    if not data_rows:
        raise ValueError('No data rows found below the header.')

    return col_map, data_rows, header_idx + 1


def _row_to_payload(row: tuple[Any, ...], col_map: dict[str, int], excel_row: int) -> dict[str, Any]:
    supplier_name = str(_cell_val(row, col_map.get('supplier_name')) or '').strip()
    supplier_roll_number = str(_cell_val(row, col_map.get('supplier_roll_number')) or '').strip()
    film_type = str(_cell_val(row, col_map.get('film_type')) or '').strip()
    coating = str(_cell_val(row, col_map.get('coating')) or '').strip()

    width_mm = _decimal_or_none(_cell_val(row, col_map.get('width_mm')))
    thickness_mic = _decimal_or_none(_cell_val(row, col_map.get('thickness_mic')))
    length_mtr = _decimal_or_none(_cell_val(row, col_map.get('length_mtr')))
    gross_weight_kg = _decimal_or_none(_cell_val(row, col_map.get('gross_weight_kg')))
    net_weight_kg = _decimal_or_none(_cell_val(row, col_map.get('net_weight_kg')))
    core_raw = _cell_val(row, col_map.get('core_weight_kg'))
    core_weight_kg = _decimal_or_none(core_raw) if core_raw not in (None, '') else None

    missing = []
    if not supplier_name:
        missing.append('Supplier Name')
    if not supplier_roll_number:
        missing.append('Roll Number')
    if not film_type:
        missing.append('Film Type')
    if not coating:
        missing.append('Chemical Coating')
    if width_mm is None or width_mm <= 0:
        missing.append('Width (mm)')
    if thickness_mic is None:
        missing.append('Thickness (mic)')
    if length_mtr is None or length_mtr <= 0:
        missing.append('Length (mtr)')
    if gross_weight_kg is None or gross_weight_kg <= 0:
        missing.append('Gross Weight (kg)')
    if net_weight_kg is None or net_weight_kg <= 0:
        missing.append('Net weight (kg)')

    if missing:
        raise ValueError(f'Row {excel_row}: missing or invalid — {", ".join(missing)}')

    return {
        'excel_row': excel_row,
        'supplier_name': supplier_name[:200],
        'supplier_roll_number': supplier_roll_number[:100],
        'film_type': film_type[:50],
        'coating': coating[:50],
        'width_mm': width_mm,
        'thickness_mic': thickness_mic,
        'length_mtr': length_mtr,
        'gross_weight_kg': gross_weight_kg,
        'net_weight_kg': net_weight_kg,
        'core_weight_kg': core_weight_kg,
    }


@dataclass
class GrnBatchImportResult:
    added: list[RollGrnEntry] = field(default_factory=list)
    skipped: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def added_count(self) -> int:
        return len(self.added)

    @property
    def skipped_count(self) -> int:
        return len(self.skipped)


def import_grn_batch_excel(file_bytes: bytes, *, created_by_id: int | None) -> GrnBatchImportResult:
    """Import rows from Excel; skip supplier+roll already in DB; assign new batch numbers only to new rows."""
    result = GrnBatchImportResult()
    _, load_workbook, _ = _openpyxl()
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f'Could not read Excel file: {exc}') from exc

    try:
        col_map, data_rows, first_data_row = _parse_rows_from_workbook(wb)
    finally:
        wb.close()

    next_seq = _max_grn_sequence()
    seen_in_file: set[tuple[str, str]] = set()

    for offset, row in enumerate(data_rows):
        excel_row = first_data_row + offset
        try:
            payload = _row_to_payload(row, col_map, excel_row)
        except ValueError as e:
            result.errors.append(str(e))
            continue

        dedup_key = (
            payload['supplier_name'].lower(),
            payload['supplier_roll_number'].lower(),
        )
        if dedup_key in seen_in_file:
            result.errors.append(
                f'Row {excel_row}: duplicate supplier + roll number in this Excel file.'
            )
            continue

        existing = get_roll_grn_by_supplier_roll(
            payload['supplier_name'],
            payload['supplier_roll_number'],
        )
        if existing:
            seen_in_file.add(dedup_key)
            result.skipped.append({
                'excel_row': payload['excel_row'],
                'supplier_roll_number': payload['supplier_roll_number'],
                'grn_batch_number': existing.grn_number,
            })
            continue

        next_seq += 1
        batch_number = format_grn_number(next_seq)
        entry = RollGrnEntry(
            grn_number=batch_number,
            supplier_name=payload['supplier_name'],
            supplier_roll_number=payload['supplier_roll_number'],
            film_type=payload['film_type'],
            coating=payload['coating'],
            width_mm=payload['width_mm'],
            thickness_mic=payload['thickness_mic'],
            length_mtr=payload['length_mtr'],
            gross_weight_kg=payload['gross_weight_kg'],
            net_weight_kg=payload['net_weight_kg'],
            core_weight_kg=payload['core_weight_kg'],
            created_by_id=created_by_id,
        )
        db.session.add(entry)
        seen_in_file.add(dedup_key)
        result.added.append(entry)

    if result.added:
        db.session.commit()
    elif not result.errors:
        db.session.rollback()
    else:
        db.session.rollback()

    return result


def build_grn_batch_template_bytes() -> bytes:
    Workbook, _, get_column_letter = _openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = 'GRN Batch Upload'
    for col, (header, sample) in enumerate(_TEMPLATE_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=2, column=col, value=sample)
    for col in range(1, len(_TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
